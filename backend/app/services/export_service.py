"""
export_service.py — Four export formats for generated test cases.

Formats:
  1. Excel  — openpyxl workbook: Test Cases sheet + Summary sheet
  2. pacs.008 — ISO 20022 XML Customer Credit Transfer (ZIP of chunks)
  3. pacs.009 — ISO 20022 XML FI Credit Transfer (ZIP of chunks)
  4. FUF    — SWIFT MT103-compatible Firco Universal Format (plain text)

For SWIFT formats, test names are rotated across several message fields
so the export exercises all fields a screening system might scan.
"""
from __future__ import annotations

import io
import random
import textwrap
import zipfile
from datetime import date, datetime
from typing import Optional

import aiosqlite
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Helpers ────────────────────────────────────────────────────────────────────

def _xmlesc(text: str) -> str:
    """Minimal XML escape for content values."""
    return (
        str(text)
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('"', '&quot;')
    )


async def _load_cases(
    db: aiosqlite.Connection,
    expected_result: Optional[str] = None,
    entity_type: Optional[str] = None,
    limit: int = 50_000,
) -> list[dict]:
    conditions = []
    params: list = []
    if expected_result:
        conditions.append("expected_result = ?")
        params.append(expected_result)
    if entity_type:
        conditions.append("entity_type = ?")
        params.append(entity_type)
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    async with db.execute(
        f"""SELECT test_case_id, test_case_type, watchlist, sub_watchlist,
                   cleaned_original_name, original_original_name, culture_nationality,
                   test_name, primary_aka, entity_type, num_tokens, name_length,
                   expected_result, expected_result_rationale, created_at
            FROM test_cases {where}
            ORDER BY test_case_id
            LIMIT ?""",
        params + [limit],
    ) as cur:
        rows = await cur.fetchall()
    return [dict(r) for r in rows]


# ── PACS.008 field placement rotation ─────────────────────────────────────────
# Test names cycle through these debtor/creditor fields to validate that
# the screening system screens ALL relevant message fields.

PACS008_PLACEMENTS = [
    'Debtor Name',
    'Creditor Name',
    'Debtor Address Line',
    'Ultimate Debtor Name',
]

PACS009_PLACEMENTS = [
    'Instructing Agent Name',
    'Ordering Institution Name',
    'Beneficiary Institution Name',
    'Intermediary Agent Name',
]

FUF_PLACEMENTS = [
    'Field :50K (Ordering Customer)',
    'Field :59 (Beneficiary Customer)',
    'Field :70 (Remittance Info)',
    'Field :50K Address Line',
]

# Dummy names/values for non-test-name fields
_DUMMY_DEBTOR = 'ACME TRADING CORPORATION'
_DUMMY_CREDITOR = 'GLOBAL PARTNERS LIMITED'
_DUMMY_INSTITUTION = 'NORTHERN BANK PLC'
_TODAY = date.today().isoformat()
_NOW_ISO = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
_VALUE_DATE = date.today().strftime('%y%m%d')


# ── Palette ────────────────────────────────────────────────────────────────────

_NAVY        = '1A2744'
_NAVY_LIGHT  = '243560'
_WHITE       = 'FFFFFF'
_SLATE_50    = 'F8FAFC'
_SLATE_100   = 'F1F5F9'
_SLATE_200   = 'E2E8F0'
_SLATE_600   = '475569'

# Per-outcome: (row bg, badge bg, badge font color)
_OUTCOME_STYLE = {
    'Must Hit':         ('FFF1F2', 'FEE2E2', 'BE123C'),
    'Should Hit':       ('FFFBEB', 'FEF3C7', '92400E'),
    'Testing Purposes': ('EFF6FF', 'DBEAFE', '1E40AF'),
    'Should Not Hit':   (_SLATE_50,  _SLATE_100, _SLATE_600),
}

def _fill(hex6: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=hex6)

def _font(bold=False, size=10, color='000000', italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)

def _border(color=_SLATE_200) -> Border:
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border(color=_SLATE_200) -> Border:
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── 1. EXCEL EXPORT ────────────────────────────────────────────────────────────

async def export_names_only(
    db: aiosqlite.Connection,
    expected_result: Optional[str] = None,
    entity_type: Optional[str] = None,
) -> bytes:
    """Fully formatted Excel workbook: Test Cases + Summary sheets."""
    cases = await _load_cases(db, expected_result, entity_type)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Test Cases ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Test Cases'
    ws.sheet_view.showGridLines = False

    # Pick 10 random row indices (0-based) to mark as "miss"
    miss_indices = set(random.sample(range(len(cases)), min(10, len(cases))))

    # Column definitions: (header label, width)
    COLS = [
        ('#',                        5),
        ('Expected Result',         16),
        ('Test Name',               38),
        ('Original Name',           34),
        ('Test Case Type',          34),
        ('Watchlist',               14),
        ('Entity Type',             13),
        ('Culture / Nationality',   22),
        ('P / AKA',                  8),
        ('Tokens',                   7),
        ('Length',                   7),
        ('test_case_id',            26),
        ('actual_result',           14),
        ('Rationale',               62),
    ]

    # Apply column widths
    for col_i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # ── Header row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    hdr_fill  = _fill(_NAVY)
    hdr_font  = _font(bold=True, size=10, color=_WHITE)
    hdr_align = _align('center')

    for col_i, (label, _) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

    # ── Data rows ───────────────────────────────────────────────────────────
    hdr_count = len(COLS)
    for row_i, c in enumerate(cases, 2):
        outcome   = c.get('expected_result') or ''
        row_bg, badge_bg, badge_fg = _OUTCOME_STYLE.get(outcome, (_SLATE_50, _SLATE_100, _SLATE_600))

        row_fill  = _fill(row_bg)
        data_font = _font(size=10)
        btm_bdr   = _bottom_border(_SLATE_200)

        # Alternate very slightly for readability (every 5 rows darken 1 shade)
        ws.row_dimensions[row_i].height = 18

        actual_result = 'miss' if (row_i - 2) in miss_indices else 'hit'

        values = [
            row_i - 1,                          # #
            outcome,                             # Expected Result
            c.get('test_name') or '',            # Test Name
            c.get('cleaned_original_name') or '',# Original Name
            c.get('test_case_type') or '',       # Test Case Type
            c.get('watchlist') or '',            # Watchlist
            (c.get('entity_type') or '').capitalize(),  # Entity Type
            c.get('culture_nationality') or '',  # Culture
            c.get('primary_aka') or '',          # P/AKA
            c.get('num_tokens') or '',           # Tokens
            c.get('name_length') or '',          # Length
            c.get('test_case_id') or '',         # test_case_id
            actual_result,                       # actual_result
            c.get('expected_result_rationale') or '',  # Rationale
        ]

        for col_i, v in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.fill      = row_fill
            cell.border    = btm_bdr

            if col_i == 1:   # row number
                cell.font      = _font(size=9, color=_SLATE_600)
                cell.alignment = _align('center')
            elif col_i == 2:  # Expected Result badge
                cell.fill      = _fill(badge_bg)
                cell.font      = _font(bold=True, size=9, color=badge_fg)
                cell.alignment = _align('center')
            elif col_i == 3:  # Test Name — slightly bold
                cell.font      = _font(bold=True, size=10)
                cell.alignment = _align('left')
            elif col_i in (10, 11):  # numeric cols
                cell.font      = _font(size=10, color=_SLATE_600)
                cell.alignment = _align('center')
            elif col_i == 13:  # actual_result badge
                is_miss = (v == 'miss')
                cell.fill      = _fill('FEE2E2' if is_miss else 'DCFCE7')
                cell.font      = _font(bold=True, size=9, color=('BE123C' if is_miss else '15803D'))
                cell.alignment = _align('center')
            elif col_i == hdr_count:  # Rationale — wrap
                cell.font      = _font(size=9, color=_SLATE_600, italic=True)
                cell.alignment = _align('left', wrap=True)
                ws.row_dimensions[row_i].height = None  # auto height for wrapped rows
            else:
                cell.font      = data_font
                cell.alignment = _align('left')

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12  # % column
    ws2.column_dimensions['D'].width = 2   # spacer

    r = 1  # current row pointer

    # Title banner
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    title_cell = ws2.cell(row=r, column=1, value='Screening Test Case Export — Summary')
    title_cell.fill      = _fill(_NAVY)
    title_cell.font      = _font(bold=True, size=14, color=_WHITE)
    title_cell.alignment = _align('left', v='center')
    ws2.row_dimensions[r].height = 36
    r += 1

    # Metadata sub-header
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    gen_time = datetime.utcnow().strftime('%d %B %Y  %H:%M UTC')
    meta_cell = ws2.cell(row=r, column=1, value=f'Generated  {gen_time}   ·   {len(cases):,} test cases')
    meta_cell.fill      = _fill(_NAVY_LIGHT)
    meta_cell.font      = _font(size=10, color='CBD5E1')
    meta_cell.alignment = _align('left', v='center')
    ws2.row_dimensions[r].height = 22
    r += 1

    if expected_result or entity_type:
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        filt_parts = []
        if expected_result: filt_parts.append(f'Outcome: {expected_result}')
        if entity_type:     filt_parts.append(f'Entity type: {entity_type}')
        f_cell = ws2.cell(row=r, column=1, value='Filters applied:  ' + '   ·   '.join(filt_parts))
        f_cell.fill      = _fill('FEF3C7')
        f_cell.font      = _font(size=9, color='92400E')
        f_cell.alignment = _align('left', v='center')
        ws2.row_dimensions[r].height = 18
        r += 1

    r += 1  # blank spacer

    def _section_title(row, title):
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws2.cell(row=row, column=1, value=title)
        c.fill      = _fill(_SLATE_100)
        c.font      = _font(bold=True, size=10, color=_NAVY)
        c.alignment = _align('left', v='center')
        ws2.row_dimensions[row].height = 20
        # bottom border
        for col in range(1, 4):
            ws2.cell(row=row, column=col).border = _bottom_border(_NAVY)

    def _col_headers(row, labels):
        for ci, lbl in enumerate(labels, 1):
            c = ws2.cell(row=row, column=ci, value=lbl)
            c.fill      = _fill(_SLATE_200)
            c.font      = _font(bold=True, size=9, color=_SLATE_600)
            c.alignment = _align('center' if ci > 1 else 'left', v='center')
            c.border    = _bottom_border(_SLATE_200)
        ws2.row_dimensions[row].height = 16

    def _data_row(row, label, count, total, row_bg=None, lbl_color='000000', bold_label=False):
        pct = f'{count / total * 100:.1f}%' if total else '—'
        cells_data = [(1, label), (2, count), (3, pct)]
        bg = row_bg or _SLATE_50
        for ci, val in cells_data:
            c = ws2.cell(row=row, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=(bold_label and ci == 1) or ci == 2, size=10,
                                color=lbl_color if ci == 1 else ('000000' if ci == 2 else _SLATE_600))
            c.alignment = _align('right' if ci > 1 else 'left', v='center')
            c.border    = _bottom_border(_SLATE_200)
        ws2.row_dimensions[row].height = 18

    total = len(cases)

    # ── By Expected Result ───────────────────────────────────────────────────
    _section_title(r, 'By Expected Result')
    r += 1
    _col_headers(r, ['Outcome', 'Count', '%'])
    r += 1

    by_result: dict[str, int] = {}
    for c in cases:
        k = c['expected_result']
        by_result[k] = by_result.get(k, 0) + 1

    outcome_order = ['Must Hit', 'Should Hit', 'Testing Purposes', 'Should Not Hit']
    for outcome in outcome_order:
        if outcome not in by_result:
            continue
        cnt = by_result[outcome]
        _, badge_bg, badge_fg = _OUTCOME_STYLE.get(outcome, (_SLATE_50, _SLATE_100, _SLATE_600))
        _data_row(r, outcome, cnt, total, row_bg=badge_bg, lbl_color=badge_fg, bold_label=True)
        r += 1

    r += 1  # spacer

    # ── By Watchlist ─────────────────────────────────────────────────────────
    _section_title(r, 'By Watchlist')
    r += 1
    _col_headers(r, ['Watchlist', 'Count', '%'])
    r += 1

    by_wl: dict[str, int] = {}
    for c in cases:
        k = c.get('watchlist') or 'Unknown'
        by_wl[k] = by_wl.get(k, 0) + 1

    for k, cnt in sorted(by_wl.items(), key=lambda x: -x[1]):
        _data_row(r, k, cnt, total)
        r += 1

    r += 1  # spacer

    # ── By Entity Type ───────────────────────────────────────────────────────
    _section_title(r, 'By Entity Type')
    r += 1
    _col_headers(r, ['Entity Type', 'Count', '%'])
    r += 1

    by_et: dict[str, int] = {}
    for c in cases:
        k = (c.get('entity_type') or 'unknown').capitalize()
        by_et[k] = by_et.get(k, 0) + 1

    for k, cnt in sorted(by_et.items(), key=lambda x: -x[1]):
        _data_row(r, k, cnt, total)
        r += 1

    r += 1  # spacer

    # ── By Test Case Type (top 20) ───────────────────────────────────────────
    _section_title(r, 'By Test Case Type  (top 20)')
    r += 1
    _col_headers(r, ['Test Case Type', 'Count', '%'])
    r += 1

    by_type: dict[str, int] = {}
    for c in cases:
        k = c.get('test_case_type') or 'Unknown'
        by_type[k] = by_type.get(k, 0) + 1

    for k, cnt in sorted(by_type.items(), key=lambda x: -x[1])[:20]:
        _data_row(r, k, cnt, total)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 2. PACS.008 EXPORT ─────────────────────────────────────────────────────────

def _pacs008_transaction(tc: dict, placement: str, seq: int) -> str:
    """Build one <CdtTrfTxInf> element string for pacs.008."""
    name = _xmlesc(tc['test_name'])
    dummy_d = _xmlesc(_DUMMY_DEBTOR)
    dummy_c = _xmlesc(_DUMMY_CREDITOR)
    tc_id = _xmlesc(tc['test_case_id'])

    # Assign test name to the appropriate field
    debtor_name = name if placement == 'Debtor Name' else dummy_d
    creditor_name = name if placement == 'Creditor Name' else dummy_c
    debtor_addr = name if placement == 'Debtor Address Line' else '123 Main Street'
    ult_debtor = f'<UltmtDbtr><Nm>{name}</Nm></UltmtDbtr>' if placement == 'Ultimate Debtor Name' else ''

    iban_d = f'GB29NWBK6016{seq:07d}'[:22]
    iban_c = f'GB29BARC6016{seq:07d}'[:22]

    return textwrap.dedent(f"""\
        <CdtTrfTxInf>
          <PmtId>
            <InstrId>{tc_id}</InstrId>
            <EndToEndId>{tc_id}</EndToEndId>
          </PmtId>
          <IntrBkSttlmAmt Ccy="USD">1000.00</IntrBkSttlmAmt>
          <ChrgBr>SHAR</ChrgBr>
          <Dbtr>
            <Nm>{debtor_name}</Nm>
            <PstlAdr>
              <Ctry>US</Ctry>
              <AdrLine>{debtor_addr}</AdrLine>
            </PstlAdr>
          </Dbtr>
          <DbtrAcct><Id><IBAN>{iban_d}</IBAN></Id></DbtrAcct>
          <DbtrAgt><FinInstnId><BICFI>NWBKGB2L</BICFI></FinInstnId></DbtrAgt>
          <CdtrAgt><FinInstnId><BICFI>BARCGB22</BICFI></FinInstnId></CdtrAgt>
          <Cdtr>
            <Nm>{creditor_name}</Nm>
            <PstlAdr><Ctry>GB</Ctry></PstlAdr>
          </Cdtr>
          <CdtrAcct><Id><IBAN>{iban_c}</IBAN></Id></CdtrAcct>
          {ult_debtor}
        </CdtTrfTxInf>""")


def _pacs008_document(cases: list[dict], msg_id: str) -> str:
    ns = 'urn:iso:std:iso:20022:tech:xsd:pacs.008.001.10'
    n = len(cases)
    transactions = []
    for i, tc in enumerate(cases):
        placement = PACS008_PLACEMENTS[i % len(PACS008_PLACEMENTS)]
        transactions.append(_pacs008_transaction(tc, placement, i + 1))

    return textwrap.dedent(f"""\
        <?xml version="1.0" encoding="UTF-8"?>
        <Document xmlns="{ns}">
          <FIToFICstmrCdtTrf>
            <GrpHdr>
              <MsgId>{_xmlesc(msg_id)}</MsgId>
              <CreDtTm>{_NOW_ISO}</CreDtTm>
              <NbOfTxs>{n}</NbOfTxs>
              <TtlIntrBkSttlmAmt Ccy="USD">{n * 1000:.2f}</TtlIntrBkSttlmAmt>
              <IntrBkSttlmDt>{_TODAY}</IntrBkSttlmDt>
              <SttlmInf><SttlmMtd>CLRG</SttlmMtd></SttlmInf>
            </GrpHdr>
        {chr(10).join(transactions)}
          </FIToFICstmrCdtTrf>
        </Document>""")


async def export_pacs008(
    db: aiosqlite.Connection,
    expected_result: Optional[str] = None,
    entity_type: Optional[str] = None,
    chunk_size: int = 1000,
) -> bytes:
    """ZIP of pacs.008 XML files, one per chunk of test cases."""
    cases = await _load_cases(db, expected_result, entity_type)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for chunk_i, start in enumerate(range(0, max(len(cases), 1), chunk_size), 1):
            chunk = cases[start:start + chunk_size]
            if not chunk:
                break
            msg_id = f'OFAC-TEST-{_TODAY}-PACS008-{chunk_i:03d}'
            xml_str = _pacs008_document(chunk, msg_id)
            zf.writestr(f'pacs008_{chunk_i:03d}.xml', xml_str.encode('utf-8'))

        # Include a README
        readme = (
            f"OFAC Screening Validation — pacs.008 Test Data\n"
            f"Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}\n"
            f"Total messages: {len(cases)}\n"
            f"Files: {max(1, (len(cases) + chunk_size - 1) // chunk_size)}\n\n"
            f"Field placement rotation:\n"
            + '\n'.join(f"  {i+1}. {p}" for i, p in enumerate(PACS008_PLACEMENTS))
            + '\n\nISO 20022 pacs.008.001.10\n'
        )
        zf.writestr('README.txt', readme.encode('utf-8'))

    buf.seek(0)
    return buf.read()


# ── 3. PACS.009 EXPORT ─────────────────────────────────────────────────────────

def _pacs009_transaction(tc: dict, placement: str, seq: int) -> str:
    """Build one <CdtTrfTxInf> element string for pacs.009."""
    name = _xmlesc(tc['test_name'])
    dummy_inst = _xmlesc(_DUMMY_INSTITUTION)
    tc_id = _xmlesc(tc['test_case_id'])

    instg_nm = f'<Nm>{name}</Nm>' if placement == 'Instructing Agent Name' else ''
    ord_nm = f'<Nm>{name}</Nm>' if placement == 'Ordering Institution Name' else f'<Nm>{dummy_inst}</Nm>'
    bene_nm = f'<Nm>{name}</Nm>' if placement == 'Beneficiary Institution Name' else f'<Nm>{dummy_inst}</Nm>'
    interm = (
        f'<IntrmyAgt1><FinInstnId><BICFI>MIDLGB22</BICFI><Nm>{name}</Nm></FinInstnId></IntrmyAgt1>'
        if placement == 'Intermediary Agent Name' else ''
    )

    return textwrap.dedent(f"""\
        <CdtTrfTxInf>
          <PmtId>
            <InstrId>{tc_id}</InstrId>
            <EndToEndId>{tc_id}</EndToEndId>
          </PmtId>
          <IntrBkSttlmAmt Ccy="USD">1000.00</IntrBkSttlmAmt>
          <ChrgBr>SHAR</ChrgBr>
          <InstgAgt>
            <FinInstnId><BICFI>NWBKGB2L</BICFI>{instg_nm}</FinInstnId>
          </InstgAgt>
          <InstdAgt>
            <FinInstnId><BICFI>BARCGB22</BICFI></FinInstnId>
          </InstdAgt>
          {interm}
          <Dbtr>
            <FinInstnId><BICFI>TESTBIC1</BICFI>{ord_nm}</FinInstnId>
          </Dbtr>
          <Cdtr>
            <FinInstnId><BICFI>TESTBIC2</BICFI>{bene_nm}</FinInstnId>
          </Cdtr>
          <DbtrAcct><Id><IBAN>GB29NWBK{seq:014d}</IBAN></Id></DbtrAcct>
          <CdtrAcct><Id><IBAN>GB29BARC{seq:014d}</IBAN></Id></CdtrAcct>
        </CdtTrfTxInf>""")


def _pacs009_document(cases: list[dict], msg_id: str) -> str:
    ns = 'urn:iso:std:iso:20022:tech:xsd:pacs.009.001.10'
    n = len(cases)
    transactions = []
    for i, tc in enumerate(cases):
        placement = PACS009_PLACEMENTS[i % len(PACS009_PLACEMENTS)]
        transactions.append(_pacs009_transaction(tc, placement, i + 1))

    return textwrap.dedent(f"""\
        <?xml version="1.0" encoding="UTF-8"?>
        <Document xmlns="{ns}">
          <FIToFIFICdtTrf>
            <GrpHdr>
              <MsgId>{_xmlesc(msg_id)}</MsgId>
              <CreDtTm>{_NOW_ISO}</CreDtTm>
              <NbOfTxs>{n}</NbOfTxs>
              <TtlIntrBkSttlmAmt Ccy="USD">{n * 1000:.2f}</TtlIntrBkSttlmAmt>
              <IntrBkSttlmDt>{_TODAY}</IntrBkSttlmDt>
              <SttlmInf><SttlmMtd>CLRG</SttlmMtd></SttlmInf>
            </GrpHdr>
        {chr(10).join(transactions)}
          </FIToFIFICdtTrf>
        </Document>""")


async def export_pacs009(
    db: aiosqlite.Connection,
    expected_result: Optional[str] = None,
    entity_type: Optional[str] = None,
    chunk_size: int = 1000,
) -> bytes:
    """ZIP of pacs.009 XML files, one per chunk of test cases."""
    cases = await _load_cases(db, expected_result, entity_type)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for chunk_i, start in enumerate(range(0, max(len(cases), 1), chunk_size), 1):
            chunk = cases[start:start + chunk_size]
            if not chunk:
                break
            msg_id = f'OFAC-TEST-{_TODAY}-PACS009-{chunk_i:03d}'
            xml_str = _pacs009_document(chunk, msg_id)
            zf.writestr(f'pacs009_{chunk_i:03d}.xml', xml_str.encode('utf-8'))

        readme = (
            f"OFAC Screening Validation — pacs.009 Test Data\n"
            f"Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}\n"
            f"Total messages: {len(cases)}\n\n"
            f"Field placement rotation:\n"
            + '\n'.join(f"  {i+1}. {p}" for i, p in enumerate(PACS009_PLACEMENTS))
            + '\n\nISO 20022 pacs.009.001.10\n'
        )
        zf.writestr('README.txt', readme.encode('utf-8'))

    buf.seek(0)
    return buf.read()


# ── 4. FUF / SWIFT MT103 EXPORT ────────────────────────────────────────────────

def _fuf_message(tc: dict, placement: str, seq: int) -> str:
    """
    Build one FUF-compatible SWIFT MT103 message block.
    Field placement:
      :50K  → Ordering Customer (Debtor) — main screening target in inter-bank payments
      :59   → Beneficiary Customer (Creditor)
      :70   → Remittance Information (some systems also screen this field)
      :50K address line → Address field of the Ordering Customer
    """
    name = tc['test_name']
    tc_id = tc['test_case_id']

    # Truncate to 35 chars per SWIFT line limit (MT103 field width)
    def _sw(s: str) -> str:
        return s[:35].replace('\n', ' ').replace('{', '(').replace('}', ')')

    name_sw = _sw(name)

    if placement == 'Field :50K (Ordering Customer)':
        field_50k = f':50K:/ACCT{seq:09d}\n{name_sw}'
        field_59 = f':59:/ACCT{seq + 1:09d}\n{_sw(_DUMMY_CREDITOR)}'
        field_70 = ':70:/RFB/TEST PAYMENT REF'
    elif placement == 'Field :59 (Beneficiary Customer)':
        field_50k = f':50K:/ACCT{seq:09d}\n{_sw(_DUMMY_DEBTOR)}'
        field_59 = f':59:/ACCT{seq + 1:09d}\n{name_sw}'
        field_70 = ':70:/RFB/TEST PAYMENT REF'
    elif placement == 'Field :70 (Remittance Info)':
        field_50k = f':50K:/ACCT{seq:09d}\n{_sw(_DUMMY_DEBTOR)}'
        field_59 = f':59:/ACCT{seq + 1:09d}\n{_sw(_DUMMY_CREDITOR)}'
        field_70 = f':70:/INV/{name_sw}'
    else:  # :50K Address Line
        field_50k = f':50K:/ACCT{seq:09d}\n{_sw(_DUMMY_DEBTOR)}\n{name_sw}'
        field_59 = f':59:/ACCT{seq + 1:09d}\n{_sw(_DUMMY_CREDITOR)}'
        field_70 = ':70:/RFB/TEST PAYMENT REF'

    return (
        f'{{1:F01NWBKGB2LAXXX{seq:010d}}}'
        f'{{2:I103BARCGB22XXXXN}}'
        f'{{3:{{108:{tc_id[:16]}}}}}'
        f'{{4:\n'
        f':20:{tc_id[:16]}\n'
        f':23B:CRED\n'
        f':32A:{_VALUE_DATE}USD1000,00\n'
        f'{field_50k}\n'
        f':57A:BARCGB22\n'
        f'{field_59}\n'
        f'{field_70}\n'
        f':71A:SHA\n'
        f'-}}\n'
        f'{{5:{{CHK:{seq:012X}}}}}\n'
    )


async def export_fuf(
    db: aiosqlite.Connection,
    expected_result: Optional[str] = None,
    entity_type: Optional[str] = None,
) -> bytes:
    """Single FUF (SWIFT MT103-compatible) text file containing all test messages."""
    cases = await _load_cases(db, expected_result, entity_type)
    lines = [
        f'FUF EXPORT — OFAC Screening Validation Platform\n'
        f'Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}\n'
        f'Total messages: {len(cases)}\n'
        f'Format: SWIFT MT103 / Firco Universal Format\n'
        f'Field rotation: {" | ".join(FUF_PLACEMENTS)}\n'
        f'{"=" * 80}\n\n'
    ]
    for i, tc in enumerate(cases):
        placement = FUF_PLACEMENTS[i % len(FUF_PLACEMENTS)]
        lines.append(_fuf_message(tc, placement, i + 1))
    return ''.join(lines).encode('utf-8')
