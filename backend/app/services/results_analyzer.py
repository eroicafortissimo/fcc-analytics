"""
results_analyzer.py — Upload, join, and compute confusion matrix statistics.

Flow:
  1. ingest_results(): parse CSV/Excel upload, join on test_case_id, upsert into
     screening_results table.
  2. compute_summary(): TP/FP/TN/FN + detection rate, FPR, precision, recall, F1
  3. compute_breakdown(by): same metrics grouped by a dimension (entity_type,
     watchlist, test_case_type, token_count, name_length_bucket, nationality)
"""
from __future__ import annotations

import csv
import io
from typing import Optional

import aiosqlite
import openpyxl


# ── Helpers ────────────────────────────────────────────────────────────────────

def _metrics(tp: int, fp: int, tn: int, fn: int) -> dict:
    """Compute derived metrics from raw confusion counts."""
    total = tp + fp + tn + fn
    detection_rate = tp / (tp + fn) if (tp + fn) > 0 else None
    fpr = fp / (fp + tn) if (fp + tn) > 0 else None
    precision = tp / (tp + fp) if (tp + fp) > 0 else None
    recall = detection_rate
    f1 = (
        2 * precision * recall / (precision + recall)
        if precision is not None and recall is not None and (precision + recall) > 0
        else None
    )

    def _r(v):
        return round(v, 4) if v is not None else None

    return {
        'total': total,
        'tp': tp, 'fp': fp, 'tn': tn, 'fn': fn,
        'detection_rate': _r(detection_rate),
        'false_positive_rate': _r(fpr),
        'precision': _r(precision),
        'recall': _r(recall),
        'f1': _r(f1),
    }


def _parse_csv_or_excel(content: bytes, filename: str) -> list[dict]:
    """Parse an uploaded CSV or Excel file into a list of normalised dicts."""
    fname = (filename or '').lower()
    if fname.endswith('.xlsx') or fname.endswith('.xls'):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            return []
        headers = [str(h or '').lower().strip().replace(' ', '_').replace('/', '_') for h in raw_headers]
        return [
            {k: str(v or '').strip() for k, v in zip(headers, row)}
            for row in rows_iter
        ]
    else:
        # CSV (UTF-8 with optional BOM)
        text = content.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        return [
            {k.lower().strip().replace(' ', '_'): v.strip() for k, v in row.items()}
            for row in reader
        ]


def _normalise_result(raw: str) -> Optional[str]:
    v = raw.upper().strip()
    if v in ('HIT', 'YES', '1', 'TRUE', 'MATCH', 'ALERT'):
        return 'HIT'
    if v in ('MISS', 'NO', '0', 'FALSE', 'NO_MATCH', 'NO MATCH', 'NOMATCH'):
        return 'MISS'
    return None


def _normalise_expected(raw: str) -> str:
    """Convert test case expected_result labels (Must Hit / Should Not Hit) to HIT / MISS."""
    v = raw.strip()
    if v in ('Must Hit', 'Should Hit', 'HIT'):
        return 'HIT'
    if v in ('Should Not Hit', 'MISS'):
        return 'MISS'
    return 'MISS'  # 'Testing Purposes' etc.


# ── Core service functions ─────────────────────────────────────────────────────

async def ingest_results(file, db: aiosqlite.Connection) -> dict:
    """
    Parse an uploaded CSV/Excel results file and upsert into screening_results.

    Required columns (case-insensitive, spaces OK):
      test_case_id  — matches test_cases.test_case_id
      actual_result — HIT / MISS  (also accepts yes/no/1/0/alert/match)

    Optional columns:
      match_score, matched_list_entry, alert_details
    """
    content = await file.read()
    rows = _parse_csv_or_excel(content, file.filename or '')

    # Flexible column aliases
    _TC_ID_KEYS = ('test_case_id', 'id', 'tcid', 'tc_id', 'case_id')
    _ACTUAL_KEYS = ('actual_result', 'actual', 'result', 'screening_result',
                    'system_result', 'hit_miss', 'outcome')
    _NAME_KEYS = ('test_name', 'name', 'screened_name', 'party_name',
                  'customer_name', 'entity_name', 'subject_name')

    def _get(row, keys):
        for k in keys:
            v = row.get(k, '').strip()
            if v:
                return v
        return ''

    stats = {'total_rows': len(rows), 'matched': 0, 'unmatched': 0,
             'skipped_bad_result': 0, 'matched_by_name': 0}
    to_upsert: list[tuple] = []

    for row in rows:
        tc_id = _get(row, _TC_ID_KEYS)
        actual_raw = _get(row, _ACTUAL_KEYS)

        actual = _normalise_result(actual_raw)
        if actual is None:
            stats['skipped_bad_result'] += 1
            continue

        tc_row = None

        # Primary: match by test_case_id
        if tc_id:
            async with db.execute(
                "SELECT test_case_id, expected_result, test_name FROM test_cases WHERE test_case_id = ?",
                (tc_id,),
            ) as cur:
                tc_row = await cur.fetchone()

        # Fallback: match by test_name (case-insensitive) when tc_id absent or not found
        if tc_row is None:
            name_raw = _get(row, _NAME_KEYS)
            if name_raw:
                async with db.execute(
                    "SELECT test_case_id, expected_result, test_name FROM test_cases WHERE test_name = ? COLLATE NOCASE LIMIT 1",
                    (name_raw,),
                ) as cur:
                    tc_row = await cur.fetchone()
                if tc_row:
                    tc_id = tc_row[0]
                    stats['matched_by_name'] += 1

        if not tc_row:
            stats['unmatched'] += 1
            continue

        expected, test_name = _normalise_expected(tc_row[1]), tc_row[2]

        score = _get(row, ('match_score', 'score', 'similarity', 'confidence'))
        try:
            score_val: Optional[float] = float(score) if score else None
        except ValueError:
            score_val = None

        to_upsert.append((
            tc_id,
            test_name,
            expected,
            actual,
            score_val,
            _get(row, ('matched_list_entry', 'matched_entry', 'list_entry', 'matched_name')),
            _get(row, ('alert_details', 'details', 'notes', 'comments')),
        ))
        stats['matched'] += 1

    if to_upsert:
        await db.executemany(
            """INSERT INTO screening_results
               (test_case_id, test_name, expected_result, actual_result,
                match_score, matched_list_entry, alert_details)
               VALUES (?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(test_case_id) DO UPDATE SET
                   actual_result=excluded.actual_result,
                   match_score=excluded.match_score,
                   matched_list_entry=excluded.matched_list_entry,
                   alert_details=excluded.alert_details,
                   uploaded_at=datetime('now')""",
            to_upsert,
        )
        await db.commit()

    stats['inserted'] = len(to_upsert)
    return stats


async def compute_summary(db: aiosqlite.Connection) -> dict:
    """Return overall confusion matrix and derived metrics."""
    async with db.execute("""
        SELECT
            SUM(CASE WHEN expected_result='HIT' AND actual_result='HIT'  THEN 1 ELSE 0 END),
            SUM(CASE WHEN expected_result='MISS' AND actual_result='HIT'  THEN 1 ELSE 0 END),
            SUM(CASE WHEN expected_result='MISS' AND actual_result='MISS' THEN 1 ELSE 0 END),
            SUM(CASE WHEN expected_result='HIT' AND actual_result='MISS' THEN 1 ELSE 0 END),
            COUNT(*)
        FROM screening_results
    """) as cur:
        row = await cur.fetchone()

    tp, fp, tn, fn = (row[0] or 0), (row[1] or 0), (row[2] or 0), (row[3] or 0)
    result = _metrics(tp, fp, tn, fn)
    result['total'] = row[4] or 0
    return result


async def compute_breakdown(by: str, db: aiosqlite.Connection) -> list[dict]:
    """
    Return per-group confusion metrics.

    by values:
      entity_type, watchlist, culture_nationality, num_tokens, name_length_bucket,
      test_case_type
    """
    # Validate dimension
    valid_dims = {
        'entity_type': 'tc.entity_type',
        'watchlist': 'tc.watchlist',
        'culture_nationality': "COALESCE(tc.culture_nationality, 'Unknown')",
        'num_tokens': 'CAST(tc.num_tokens AS TEXT)',
        'test_case_type': 'tc.test_case_type',
        'name_length_bucket': (
            "CASE "
            "WHEN tc.name_length BETWEEN 1 AND 4 THEN '1–4'"
            " WHEN tc.name_length BETWEEN 5 AND 9 THEN '5–9'"
            " WHEN tc.name_length BETWEEN 10 AND 14 THEN '10–14'"
            " WHEN tc.name_length BETWEEN 15 AND 19 THEN '15–19'"
            " WHEN tc.name_length BETWEEN 20 AND 24 THEN '20–24'"
            " ELSE '25+' END"
        ),
    }
    dim_expr = valid_dims.get(by, 'tc.entity_type')

    async with db.execute(f"""
        SELECT
            {dim_expr} AS dim_value,
            SUM(CASE WHEN sr.expected_result='HIT' AND sr.actual_result='HIT'  THEN 1 ELSE 0 END) AS tp,
            SUM(CASE WHEN sr.expected_result='MISS' AND sr.actual_result='HIT'  THEN 1 ELSE 0 END) AS fp,
            SUM(CASE WHEN sr.expected_result='MISS' AND sr.actual_result='MISS' THEN 1 ELSE 0 END) AS tn,
            SUM(CASE WHEN sr.expected_result='HIT' AND sr.actual_result='MISS' THEN 1 ELSE 0 END) AS fn
        FROM screening_results sr
        JOIN test_cases tc ON sr.test_case_id = tc.test_case_id
        GROUP BY dim_value
        ORDER BY (tp + fn) DESC
    """) as cur:
        rows = await cur.fetchall()

    return [
        {'dimension': row[0], **_metrics(row[1] or 0, row[2] or 0, row[3] or 0, row[4] or 0)}
        for row in rows
    ]


async def get_results_table(
    db: aiosqlite.Connection,
    page: int = 1,
    page_size: int = 100,
    outcome_filter: Optional[str] = None,
    entity_type: Optional[str] = None,
    search: Optional[str] = None,
) -> dict:
    """Return paginated screening_results joined with test_cases."""
    conditions = []
    params: list = []

    if outcome_filter == 'FN':
        conditions.append("sr.expected_result='HIT' AND sr.actual_result='MISS'")
    elif outcome_filter == 'FP':
        conditions.append("sr.expected_result='MISS' AND sr.actual_result='HIT'")
    elif outcome_filter == 'TP':
        conditions.append("sr.expected_result='HIT' AND sr.actual_result='HIT'")
    elif outcome_filter == 'TN':
        conditions.append("sr.expected_result='MISS' AND sr.actual_result='MISS'")

    if entity_type:
        conditions.append("tc.entity_type = ?")
        params.append(entity_type)

    if search:
        conditions.append("(sr.test_name LIKE ? OR tc.cleaned_original_name LIKE ?)")
        like = f'%{search}%'
        params.extend([like, like])

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    offset = (page - 1) * page_size

    async with db.execute(
        f"""SELECT COUNT(*) FROM screening_results sr
            JOIN test_cases tc ON sr.test_case_id = tc.test_case_id {where}""",
        params,
    ) as cur:
        total = (await cur.fetchone())[0]

    async with db.execute(
        f"""SELECT sr.test_case_id, sr.test_name, tc.cleaned_original_name,
                   tc.test_case_type, tc.entity_type, tc.watchlist,
                   sr.expected_result, sr.actual_result,
                   sr.match_score, sr.matched_list_entry, sr.alert_details
            FROM screening_results sr
            JOIN test_cases tc ON sr.test_case_id = tc.test_case_id
            {where}
            ORDER BY sr.test_case_id
            LIMIT ? OFFSET ?""",
        params + [page_size, offset],
    ) as cur:
        rows = await cur.fetchall()

    return {
        'total': total,
        'page': page,
        'page_size': page_size,
        'items': [dict(r) for r in rows],
    }


async def export_results_excel(db: aiosqlite.Connection) -> bytes:
    """Export full results table as a styled Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    summary = await compute_summary(db)

    async with db.execute("""
        SELECT sr.test_case_id, sr.test_name, tc.cleaned_original_name,
               tc.test_case_type, tc.entity_type, tc.watchlist, tc.culture_nationality,
               sr.expected_result, sr.actual_result,
               sr.match_score, sr.matched_list_entry, sr.alert_details
        FROM screening_results sr
        JOIN test_cases tc ON sr.test_case_id = tc.test_case_id
        ORDER BY sr.test_case_id
    """) as cur:
        rows = await cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'

    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)

    headers = [
        'Test Case ID', 'Test Name', 'Original Name', 'Test Case Type',
        'Entity Type', 'Watchlist', 'Nationality', 'Expected', 'Actual',
        'Match Score', 'Matched Entry', 'Alert Details',
    ]
    col_widths = [26, 32, 32, 38, 12, 14, 18, 10, 10, 12, 30, 30]
    outcome_fills = {
        ('HIT', 'HIT'): PatternFill(fill_type='solid', fgColor='D1FAE5'),   # TP green
        ('MISS', 'MISS'): PatternFill(fill_type='solid', fgColor='DBEAFE'), # TN blue
        ('HIT', 'MISS'): PatternFill(fill_type='solid', fgColor='FEE2E2'),  # FN red
        ('MISS', 'HIT'): PatternFill(fill_type='solid', fgColor='FEF9C3'),  # FP yellow
    }

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'

    for row_i, r in enumerate(rows, 2):
        row_dict = dict(r)
        values = [
            row_dict.get('test_case_id'), row_dict.get('test_name'),
            row_dict.get('cleaned_original_name'), row_dict.get('test_case_type'),
            row_dict.get('entity_type'), row_dict.get('watchlist'),
            row_dict.get('culture_nationality'),
            row_dict.get('expected_result'), row_dict.get('actual_result'),
            row_dict.get('match_score'), row_dict.get('matched_list_entry'),
            row_dict.get('alert_details'),
        ]
        fill = outcome_fills.get(
            (row_dict.get('expected_result'), row_dict.get('actual_result'))
        )
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill and col in (8, 9):
                cell.fill = fill

    # Summary sheet
    ws2 = wb.create_sheet('Confusion Matrix')
    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16

    def _h(row, col, val, colour='1E3A5F'):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(fill_type='solid', fgColor=colour)
        cell.alignment = Alignment(horizontal='center')
        return cell

    def _v(row, col, val):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(horizontal='center')
        return cell

    ws2.cell(row=1, column=1, value=f'Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}')
    _h(3, 2, 'Predicted HIT', '15803D')
    _h(3, 3, 'Predicted MISS', 'B91C1C')
    _h(4, 1, 'Expected HIT', '15803D')
    _v(4, 2, summary['tp']).fill = PatternFill(fill_type='solid', fgColor='D1FAE5')
    _v(4, 3, summary['fn']).fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
    _h(5, 1, 'Expected MISS', 'B91C1C')
    _v(5, 2, summary['fp']).fill = PatternFill(fill_type='solid', fgColor='FEF9C3')
    _v(5, 3, summary['tn']).fill = PatternFill(fill_type='solid', fgColor='DBEAFE')

    metrics = [
        ('Total Results', summary['total']),
        ('Detection Rate (Recall)', f"{(summary['detection_rate'] or 0)*100:.1f}%"),
        ('False Positive Rate', f"{(summary['false_positive_rate'] or 0)*100:.1f}%"),
        ('Precision', f"{(summary['precision'] or 0)*100:.1f}%"),
        ('F1 Score', f"{(summary['f1'] or 0)*100:.1f}%"),
    ]
    for ri, (label, value) in enumerate(metrics, 7):
        ws2.cell(row=ri, column=1, value=label)
        ws2.cell(row=ri, column=2, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
